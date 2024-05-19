import os
import time
from openpyxl import load_workbook
from selenium import webdriver

from selenium.webdriver.common.by import By

from selenium.webdriver.support.wait import WebDriverWait

from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.edge.options import Options as EdgeOptions


option = EdgeOptions()
option.add_argument("start-maximized")
driver = webdriver.Edge(options = option)



timeOut = 55
workbook = load_workbook(filename = "PythonINCMaker.xlsx")
sheet = workbook.active

if sheet["A2"].value != None:
    
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=1000, max_col=1):

        driver.get("#############")     
        WebDriverWait(driver, timeOut).until(EC.title_contains(" | Incident | ServiceNow"))
        os.system('cls')
        for cell in row:
            
            shadow_section = driver.execute_script('''return document.querySelector("body > macroponent-f51912f4c700201072b211d4d8c26010").shadowRoot.querySelector("div > sn-canvas-appshell-root > sn-canvas-appshell-layout")''')
            iframe = shadow_section.find_element(By.TAG_NAME, 'iframe')
            driver.switch_to.frame (iframe)
            if cell != None:

                for cell in row: #cell will be the current row and so the ticket

                    try: #caller is the user that is having the issue
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[3]/div[2]/div[2]/input')))
                        callerField = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[3]/div[2]/div[2]/input')
                        callerField.send_keys(cell.value)   

                    except TimeoutException:
                        print ("Snow failed, run script again...")
                        print("Is the user EID load correctly?")
                    time.sleep(2)

                    try: #CLASS BUTTON OPEN THE CLASSIFICATION WINDOW
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[8]/div[2]/div[2]/span[2]/button')))
                        expandClassificationsButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[8]/div[2]/div[2]/span[2]/button')
                        expandClassificationsButton.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")
                    time.sleep(1)

                    ########## edge window switch
                    original_window = driver.current_window_handle  #
                    incTypesWindow = driver.window_handles[1] ## from window 0 to window 1 in the windows list. 
                    driver.switch_to.window(incTypesWindow)   ## 
                    ########## 
                    try: #CLASSIFY ROOT
                        cellRow = str(cell.row)  
                        classRoot = sheet["c"+cellRow].value
                        classRooToConc = str(classRoot)
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+classRooToConc+']/table')))
                        classRootButton = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+classRooToConc+']/table')
                        classRootButton.click()

                    except TimeoutException:
                        print ("Snow failed, run script again...")



                    ########## ########## 
                    driver.switch_to.window(original_window) #vuelve al iframe 1er ciclo
                    shadow_section = driver.execute_script('''return document.querySelector("body > macroponent-f51912f4c700201072b211d4d8c26010").shadowRoot.querySelector("div > sn-canvas-appshell-root > sn-canvas-appshell-layout")''')
                    iframe = shadow_section.find_element(By.TAG_NAME, 'iframe')
                    driver.switch_to.frame (iframe)
                    ########## 

                    try: #CLASS BUTTON OPEN THE CLASSIFICATION WINDOW
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[8]/div[2]/div[2]/span[2]/button')))
                        expandClassificationsButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[8]/div[2]/div[2]/span[2]/button')
                        expandClassificationsButton.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")
                    ########## 
                    original_window = driver.current_window_handle #PASA ALA VENTANA 2DO CICLO
                    incTypesWindow = driver.window_handles[1]
                    driver.switch_to.window(incTypesWindow)
                    ########## 
                    try: #SUB CLASSIFY
                        cellRow = str(cell.row)
                        subClass = sheet["d"+cellRow].value
                        subClassToConc = str(subClass)
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+classRooToConc+']/div/div['+subClassToConc+']/table/tbody/tr/td[3]/a')))
                        subClasstButton = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+classRooToConc+']/div/div['+subClassToConc+']/table/tbody/tr/td[3]/a')
                        subClasstButton.click()

                    except TimeoutException:
                        print ("Snow failed, run script again...")
                    
                    driver.switch_to.window(original_window)
                    shadow_section = driver.execute_script('''return document.querySelector("body > macroponent-f51912f4c700201072b211d4d8c26010").shadowRoot.querySelector("div > sn-canvas-appshell-root > sn-canvas-appshell-layout")''')
                    iframe = shadow_section.find_element(By.TAG_NAME, 'iframe')
                    driver.switch_to.frame (iframe)
                    time.sleep(1)
                    ########## ########## 


                    if classRooToConc == '1': #sw class will get the app from the excel. 
                        try:
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[12]/div[2]/div[2]/input')))
                            appAffected = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[12]/div[2]/div[2]/input')
                            appAffected.send_keys(sheet['g'+str(cell.row)].value)
                        except TimeoutException:
                            print ("Snow failed, run script again...")

                    elif classRooToConc == '8' and subClassToConc == '14' : #compliance class will get the cmplnc item from the excel
                        try: 
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[9]/div[2]/div[2]/input')))
                            itemUncompliance = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[9]/div[2]/div[2]/input')
                            itemUncompliance.send_keys(sheet['o'+str(cell.row)].value)
                            itemUncompliance.submit()
                        except TimeoutException:
                            print ("Snow failed, run script again...")

                    elif classRooToConc == '9': #ntwk item from the excel
                        try: #ITEM NETWORK
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[9]/div[2]/div[2]/input')))
                            networkItem = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[9]/div[2]/div[2]/input')
                            networkItem.send_keys(sheet['m'+str(cell.row)].value)
                            networkItem.submit()
                        except TimeoutException:
                            print ("Snow failed, run script again...")


                    else:  
                        try:# any else just pick laptop
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[9]/div[2]/div[2]/input')))
                            itemAffected = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[9]/div[2]/div[2]/input')
                            itemAffected.send_keys('Laptop')
                        except TimeoutException:
                            print ("Snow failed, run script again...")


                    try: #always support team
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[6]/div[2]/select')))
                        channelButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[6]/div[2]/select')
                        channelButton.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")
                    try:
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[6]/div[2]/select/option[6]')))
                        channelOption = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[6]/div[2]/select/option[6]')
                        channelOption.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")

                    try: #the owner group from the excel row 'l'
                        cellRow = str(cell.row)
                        ownerGroup = sheet["l"+cellRow].value
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[10]/div[2]/div[2]/input')))
                        ownerGroupField = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[10]/div[2]/div[2]/input')
                        ownerGroupField.send_keys(ownerGroup)
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[11]/div[2]/div[2]/input')))
                        assignmentGroup = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[11]/div[2]/div[2]/input')
                        assignmentGroup.send_keys(ownerGroup)
                    except TimeoutException:
                        print ("Snow failed, run script again...")

                    #time.sleep(5)
                    
                    
                    try: #/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[11]/div[2]/div[2]/span[2]/button
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[3]/div[3]/a[2]')))
                        assetButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[3]/div[3]/a[2]')
                        assetButton.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")
                        print("The EID exists? check and run again")
                    time.sleep(2)
                    try:
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[9]/div/div/header/button')))
                        exitEmerWindow = driver.find_element(By.XPATH, '/html/body/div[9]/div/div/header/button')
                        exitEmerWindow.click()

                    except TimeoutException:
                        print ("Snow failed, run script again...")
                    time.sleep(3)

                    try:                                                                                #
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[7]/div/div/div/rendered_body/div/div/span/div/div[5]/div[1]/table/tbody/tr/td[3]/a')))
                        assetField = driver.find_element(By.XPATH, '/html/body/div[7]/div/div/div/rendered_body/div/div/span/div/div[5]/div[1]/table/tbody/tr/td[3]/a')
                        machineName = assetField.text
                    except TimeoutException:
                        print ("Snow failed, run script again...")
                    try:
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[7]/div/div/header/button')))
                        exitAssets = driver.find_element(By.XPATH, '/html/body/div[7]/div/div/header/button')
                        exitAssets.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")

                    if machineName[:3] == 'CPX' or machineName[:3] == 'C11' or machineName[:3] == 'BA5':
                        try:#this step open the assets assigned to the user and pick the first one that always is the most recently logged
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[11]/div[2]/div[2]/input')))
                            affectedCi = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[11]/div[2]/div[2]/input')
                            affectedCi.send_keys(machineName)
                        except TimeoutException:
                            print ("Snow failed, run script again...")
                    else:
                        try:#if the machine name dont start with C 
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[11]/div[2]/div[2]/input')))
                            affectedCi = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[1]/div[11]/div[2]/div[2]/input')
                            affectedCi.send_keys("Laptop")
                        except TimeoutException:
                            print ("Snow failed, run script again...")

                    try: #LTS member to assign the incident
                        cellRow = str(cell.row)
                        assignedTo = sheet["b"+cellRow].value
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[12]/div[2]/div[2]/input')))
                        assignedToField = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[12]/div[2]/div[2]/input')
                        assignedToField.send_keys(assignedTo)
                    except TimeoutException:
                        print ("Snow failed, run script again...")

                    try: #short description
                        cellRow = str(cell.row)
                        shortDesc = sheet["e"+cellRow].value
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[2]/div/div[1]/div[2]/input[3]')))
                        shortDescField = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[2]/div/div[1]/div[2]/input[3]')
                        shortDescField.send_keys(shortDesc)
                    except TimeoutException:
                        print ("Snow failed, run script again...")

                    try: #add full desc from column 'f'
                        cellRow = str(cell.row)
                        fullDesc = sheet["f"+cellRow].value
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[2]/div/div[2]/div[2]/textarea')))
                        fullDescField = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[2]/div/div[2]/div[2]/textarea')
                        fullDescField.send_keys(fullDesc)
                    except TimeoutException:
                        print ("Snow failed, run script again...")

                        
                    try: #ticket state 
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select')))
                        stateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select')
                        stateButton.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")

                    try:#WORK IN PROGRESS ALWAYS BEFORE  
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select/option[3]')))
                        stateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select/option[3]')
                        stateButton.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")
                    time.sleep(1)

                    stateOptionToConc = sheet['h'+str(cell.row)].value ## retrieve the state of the ticket to know if the ticket is resolved       

                    if (stateOptionToConc == 10): #if the ticket is resolved, then enter to the resolution phase            
                        try:   
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select/option['+str(stateOptionToConc)+']')))
                            stateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select/option['+str(stateOptionToConc)+']')
                            stateButton.click()
                        except TimeoutException:
                            print ("Snow failed, run script again...")    
            
                        driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")
                        driver.execute_script("window.scrollTo(5,document.body.scrollHeight)")

                        try: #RESOLUTION FIELD 
                            
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[2]/div[2]/div[2]/span[2]/button')))
                            resolutionButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[2]/div[2]/div[2]/span[2]/button')
                            resolutionButton.click()
                        except TimeoutException:
                            print ("Snow failed, run script again...")

                        original_window = driver.current_window_handle 
                        incTypesWindow = driver.window_handles[1]
                        driver.switch_to.window(incTypesWindow)

                        try: #RESOLUTION ROOT
                            resolutionRoot = sheet['i'+str(cell.row)].value

                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+str(resolutionRoot)+']/table/tbody/tr/td[3]/a')))
                            resolutionRoot = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+str(resolutionRoot)+']/table/tbody/tr/td[3]/a')
                            resolutionRoot.click()
                        except TimeoutException:
                            print ("Snow failed, run script again...") 
                        
                        driver.switch_to.window(original_window)
                        shadow_section = driver.execute_script('''return document.querySelector("body > macroponent-f51912f4c700201072b211d4d8c26010").shadowRoot.querySelector("div > sn-canvas-appshell-root > sn-canvas-appshell-layout")''')
                        iframe = shadow_section.find_element(By.TAG_NAME, 'iframe')
                        driver.switch_to.frame (iframe)
                        time.sleep(2)

                        try: #sUB RESOLUTION FIELD
                            
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[2]/div[2]/div[2]/span[2]/button')))
                            resolutionButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[2]/div[2]/div[2]/span[2]/button')
                            resolutionButton.click()
                        except TimeoutException:
                            print ("Snow failed, run script again...")

                        original_window = driver.current_window_handle #PASA ALA VENTANA 2DO CICLO
                        incTypesWindow = driver.window_handles[1]
                        driver.switch_to.window(incTypesWindow)
                        time.sleep(1)
                        

                        try: #SUB RESOLUTION FIELD
                            subResolutionToConc = sheet['j'+str(cell.row)].value
                            resolutionRoot = sheet['i'+str(cell.row)].value
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+str(resolutionRoot)+']/div/div['+str(subResolutionToConc)+']/table/tbody/tr/td[3]/a')))
                            subResolution = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[3]/div/div/div/div/div['+str(resolutionRoot)+']/div/div['+str(subResolutionToConc)+']/table/tbody/tr/td[3]/a')                                 
                            subResolution.click()                                                
                        except TimeoutException:
                            print ("Snow failed, run script again...") 
                        
                        driver.switch_to.window(original_window)
                        shadow_section = driver.execute_script('''return document.querySelector("body > macroponent-f51912f4c700201072b211d4d8c26010").shadowRoot.querySelector("div > sn-canvas-appshell-root > sn-canvas-appshell-layout")''')
                        iframe = shadow_section.find_element(By.TAG_NAME, 'iframe')
                        driver.switch_to.frame (iframe)
                        time.sleep(1)
    
                        partFixed = [2, 3, 4, 7, 8, 9, 11]
                        if  (resolutionRoot == 3) and (subResolutionToConc in partFixed):
                            #CICLO 3
                            try: #PART FIXED FIELD
                            
                                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[2]/div[2]/div[2]/span[2]/button')))
                                        resolutionButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[2]/div[2]/div[2]/span[2]/button')
                                        resolutionButton.click()
                            except TimeoutException:
                                print ("Snow failed, run script again...")

                            original_window = driver.current_window_handle #PASA ALA VENTANA 3ER CICLO
                            incTypesWindow = driver.window_handles[1]
                            driver.switch_to.window(incTypesWindow)
                            time.sleep(3)


                            try: #PART FIXED FIELD
                                partToFix = sheet['k'+str(cell.row)].value
                                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH,'/html/body/div[1]/div/div[3]/div/div/div/div/div[3]/div/div['+str(subResolutionToConc)+']/div/div['+str(partToFix)+']/table/tbody/tr/td[3]/a')))
                                fixedPart = driver.find_element(By.XPATH,'/html/body/div[1]/div/div[3]/div/div/div/div/div[3]/div/div['+str(subResolutionToConc)+']/div/div['+str(partToFix)+']/table/tbody/tr/td[3]/a')
                                fixedPart.click()
                            except TimeoutException:
                                print ("Snow failed, run script again...")

                            driver.switch_to.window(original_window)
                            shadow_section = driver.execute_script('''return document.querySelector("body > macroponent-f51912f4c700201072b211d4d8c26010").shadowRoot.querySelector("div > sn-canvas-appshell-root > sn-canvas-appshell-layout")''')
                            iframe = shadow_section.find_element(By.TAG_NAME, 'iframe')
                            driver.switch_to.frame (iframe)
                            time.sleep(3)

                        try: #CLOSE CODE
                                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[8]/div[2]/select')))
                                closeCodeButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[8]/div[2]/select')
                                closeCodeButton.click()
                        except TimeoutException:
                                print ("Snow failed, run script again...")

                        try: #CLOSE CODE SOLVED PERMANENTLY
                                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[8]/div[2]/select/option[3]')))
                                closeCodeButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[1]/div[1]/div[8]/div[2]/select/option[3]')
                                closeCodeButton.click()
                        except TimeoutException:
                                print ("Snow failed, run script again...")



                        try: #CLOSE NOTES
                                closeNotes =  sheet['n'+str(cell.row)].value
                                WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[2]/div/div/div[2]/textarea')))
                                closeCodeArea = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[5]/span/div/div[2]/div/div/div[2]/textarea')
                                closeCodeArea.send_keys(closeNotes)
                        except TimeoutException:
                                print ("Snow failed, run script again...")            
                        
                    else:
                        try:   
                            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select/option['+str(stateOptionToConc)+']')))
                            stateButton = driver.find_element(By.XPATH, '/html/body/div[2]/form/span[1]/span/div[5]/div[1]/div[2]/div[7]/div[2]/select/option['+str(stateOptionToConc)+']')
                            stateButton.click()
                        except TimeoutException:
                            print ("Snow failed, run script again...") 
                    try:
                        WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/span/span/nav/div/div[2]/span[1]/span[2]/span/button[1]')))
                        submitButton = driver.find_element(By.XPATH, '/html/body/div[1]/span/span/nav/div/div[2]/span[1]/span[2]/span/button[1]')
                        submitButton.click()
                    except TimeoutException:
                        print ("Snow failed, run script again...")

            
                cellRow = str(cell.row + 1)
                if sheet["A"+ cellRow] != None:
                    driver.switch_to.new_window("tab")
                    newTab = driver.current_window_handle
                    driver.switch_to.window(original_window)
                    driver.close()
                    driver.switch_to.window(newTab)

        else:
            print("TICKETS HAVE BEEN CREATED")
else:
    print("NO TICKETS TO CREATE. ADD TICKETS TO THE EXCEL FILE")

